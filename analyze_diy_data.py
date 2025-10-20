"""
DIY Kits Data Analysis Script
Analyzes the structure of the DIY Kits Master Bible Excel file
to understand tab structures, create fuzzy matching, and generate test questions
"""

import openpyxl
import pandas as pd
from rapidfuzz import fuzz, process
import json
from collections import defaultdict

# Load the workbook
EXCEL_PATH = "data/DIY Kits Master Bible for Merchandising.xlsx"
print("Loading workbook...")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# ===== TASK 1: Extract all sheet names =====
print("\n" + "="*80)
print("TASK 1: ALL SHEET NAMES")
print("="*80)

sheet_names = wb.sheetnames
print(f"\nTotal number of sheets: {len(sheet_names)}")
print("\nFirst 3 sheets (potentially different structure):")
for i, name in enumerate(sheet_names[:3], 1):
    print(f"  {i}. {name}")

print("\nRemaining sheets (product tabs):")
for i, name in enumerate(sheet_names[3:], 4):
    print(f"  {i}. {name}")

# Save all sheet names to a file
with open("output_sheet_names.json", "w") as f:
    json.dump({
        "total_sheets": len(sheet_names),
        "first_3_sheets": sheet_names[:3],
        "product_sheets": sheet_names[3:],
        "all_sheets": sheet_names
    }, f, indent=2)

print("\n✓ Saved to: output_sheet_names.json")

# ===== TASK 2: Analyze structure of first 3 sheets =====
print("\n" + "="*80)
print("TASK 2: ANALYZING FIRST 3 SHEETS STRUCTURE")
print("="*80)

first_3_analysis = {}

for sheet_name in sheet_names[:3]:
    print(f"\n--- Analyzing: {sheet_name} ---")
    sheet = wb[sheet_name]
    
    # Get dimensions
    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"Dimensions: {max_row} rows x {max_col} columns")
    
    # Get first 10 rows to understand structure
    sample_data = []
    for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
        sample_data.append(row)
        # Show first few rows
        if row_idx <= 5:
            non_empty = [cell for cell in row if cell is not None]
            if non_empty:
                print(f"  Row {row_idx}: {non_empty[:5]}{'...' if len(non_empty) > 5 else ''}")
    
    # Try to identify headers
    potential_headers = []
    for row in sample_data[:3]:
        if row and any(cell for cell in row):
            potential_headers.append([cell for cell in row if cell])
    
    first_3_analysis[sheet_name] = {
        "rows": max_row,
        "columns": max_col,
        "sample_headers": potential_headers
    }

# ===== TASK 3: Analyze structure of product sheets =====
print("\n" + "="*80)
print("TASK 3: ANALYZING PRODUCT SHEETS STRUCTURE")
print("="*80)

# Sample a few product sheets
sample_product_sheets = sheet_names[3:8] if len(sheet_names) > 8 else sheet_names[3:]
product_analysis = {}

for sheet_name in sample_product_sheets:
    print(f"\n--- Analyzing: {sheet_name} ---")
    sheet = wb[sheet_name]
    
    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"Dimensions: {max_row} rows x {max_col} columns")
    
    # Get first 15 rows
    sample_data = []
    for row_idx, row in enumerate(sheet.iter_rows(max_row=15, values_only=True), 1):
        sample_data.append(row)
        if row_idx <= 10:
            non_empty = [cell for cell in row if cell is not None]
            if non_empty:
                print(f"  Row {row_idx}: {non_empty[:5]}{'...' if len(non_empty) > 5 else ''}")
    
    product_analysis[sheet_name] = {
        "rows": max_row,
        "columns": max_col,
        "sample_data": [[cell for cell in row if cell] for row in sample_data[:10]]
    }

# Save structure analysis
structure_output = {
    "first_3_sheets": first_3_analysis,
    "sample_product_sheets": product_analysis
}

with open("output_structure_analysis.json", "w") as f:
    json.dump(structure_output, f, indent=2, default=str)

print("\n✓ Saved to: output_structure_analysis.json")

# ===== TASK 4: Fuzzy matching for sheet names =====
print("\n" + "="*80)
print("TASK 4: FUZZY MATCHING SOLUTION FOR TAB NAMES")
print("="*80)

def find_best_matching_sheet(query, sheet_names, threshold=60):
    """
    Find the best matching sheet name using fuzzy matching
    Returns: (best_match, score, all_matches)
    """
    product_sheets = [s for s in sheet_names if s not in sheet_names[:3]]
    
    # Get top 5 matches
    matches = process.extract(query, product_sheets, scorer=fuzz.token_sort_ratio, limit=5)
    
    best_match = matches[0] if matches and matches[0][1] >= threshold else None
    
    return best_match, matches

# Test fuzzy matching with various queries
test_queries = [
    "Pretty in Pink",
    "pretty pink",
    "PRETTY IN PINK",
    "Pretty n Pink",
    "Pretty & Pink",
    "pink pretty",
    "Rustic",
    "rustic charm",
    "Garden Party",
    "garden",
    "Romance",
    "romantic",
]

print("\nTesting fuzzy matching:")
fuzzy_results = {}

for query in test_queries:
    best, all_matches = find_best_matching_sheet(query, sheet_names)
    print(f"\nQuery: '{query}'")
    if best:
        print(f"  ✓ Best match: '{best[0]}' (score: {best[1]})")
    else:
        print(f"  ✗ No good match found")
    print(f"  Top 3 candidates: {all_matches[:3]}")
    
    fuzzy_results[query] = {
        "best_match": best[0] if best else None,
        "score": best[1] if best else 0,
        "top_candidates": [{"name": m[0], "score": m[1]} for m in all_matches[:3]]
    }

with open("output_fuzzy_matching_tests.json", "w") as f:
    json.dump(fuzzy_results, f, indent=2)

print("\n✓ Saved to: output_fuzzy_matching_tests.json")

# ===== TASK 5: Generate common questions bank =====
print("\n" + "="*80)
print("TASK 5: COMMON QUESTIONS BANK")
print("="*80)

common_questions = {
    "tab_lookup": [
        "Find the [Product Name] tab",
        "Which tab has information about [Product Name]?",
        "Show me the [Product Name] sheet",
        "Where can I find [Product Name]?",
        "Look up [Product Name]",
    ],
    "material_quantities": [
        "What materials are in the [Product Name] kit?",
        "How many stems are in [Product Name] [Size]?",
        "What's included in the [Product Name] [Size] kit?",
        "List all materials for [Product Name]",
        "What quantities do I need for [Product Name] [Size]?",
        "How many roses are in [Product Name] [Size]?",
    ],
    "size_specific": [
        "What's in the small [Product Name] kit?",
        "Show me the medium size for [Product Name]",
        "How many stems total in [Product Name] large?",
        "What are the sizes available for [Product Name]?",
        "Compare small vs medium [Product Name]",
    ],
    "specific_flowers": [
        "How many [Flower Type] are in [Product Name] [Size]?",
        "Does [Product Name] contain roses?",
        "What color roses are in [Product Name]?",
        "How many spray roses in [Product Name] medium?",
    ],
    "general_info": [
        "Tell me about [Product Name]",
        "What is [Product Name]?",
        "Show me everything about [Product Name]",
        "Give me details on [Product Name]",
    ],
    "comparison": [
        "What's the difference between [Product A] and [Product B]?",
        "Compare [Product A] small to [Product B] small",
        "Which has more stems, [Product A] or [Product B]?",
    ]
}

print("\nCommon question categories:")
for category, questions in common_questions.items():
    print(f"\n{category.upper().replace('_', ' ')}:")
    for q in questions[:3]:
        print(f"  - {q}")
    if len(questions) > 3:
        print(f"  ... and {len(questions) - 3} more")

with open("output_common_questions.json", "w") as f:
    json.dump(common_questions, f, indent=2)

print("\n✓ Saved to: output_common_questions.json")

# ===== TASK 6: Test scenarios =====
print("\n" + "="*80)
print("TASK 6: TEST SCENARIOS")
print("="*80)

# Create concrete test cases using actual product names
product_sheets_sample = sheet_names[3:10]

test_scenarios = []

for product in product_sheets_sample[:3]:
    test_scenarios.extend([
        {
            "test_id": f"lookup_{product.lower().replace(' ', '_')}",
            "category": "tab_lookup",
            "user_query": f"Find the {product} tab",
            "expected_behavior": f"Should find worksheet '{product}'",
            "success_criteria": f"Returns worksheet with exact or close match to '{product}'"
        },
        {
            "test_id": f"materials_{product.lower().replace(' ', '_')}_small",
            "category": "material_quantities",
            "user_query": f"What materials are in the {product} small kit?",
            "expected_behavior": f"Should read {product} sheet and return small size materials with quantities",
            "success_criteria": "Returns list of materials with quantities, mentions 'small' size"
        },
        {
            "test_id": f"stem_count_{product.lower().replace(' ', '_')}_medium",
            "category": "size_specific",
            "user_query": f"How many total stems are in {product} medium?",
            "expected_behavior": f"Should read {product} sheet, find medium size, sum all stem counts",
            "success_criteria": "Returns a numeric total stem count"
        }
    ])

# Add fuzzy matching test cases
test_scenarios.extend([
    {
        "test_id": "fuzzy_pretty_pink_lowercase",
        "category": "fuzzy_matching",
        "user_query": "find pretty in pink",
        "expected_behavior": "Should match 'Pretty in Pink' despite lowercase",
        "success_criteria": "Finds 'Pretty in Pink' worksheet"
    },
    {
        "test_id": "fuzzy_abbreviation",
        "category": "fuzzy_matching",
        "user_query": "pretty n pink",
        "expected_behavior": "Should match 'Pretty in Pink' despite different formatting",
        "success_criteria": "Finds 'Pretty in Pink' worksheet"
    }
])

# Add error cases
test_scenarios.extend([
    {
        "test_id": "nonexistent_product",
        "category": "error_handling",
        "user_query": "Find the Unicorn Rainbow Sparkle kit",
        "expected_behavior": "Should gracefully handle nonexistent product",
        "success_criteria": "Returns helpful message indicating product not found, possibly suggests similar products"
    },
    {
        "test_id": "ambiguous_query",
        "category": "error_handling",
        "user_query": "How many stems?",
        "expected_behavior": "Should ask for clarification on which product and size",
        "success_criteria": "Asks user to specify product name and/or size"
    }
])

print(f"\nGenerated {len(test_scenarios)} test scenarios:")
for i, scenario in enumerate(test_scenarios[:5], 1):
    print(f"\n{i}. {scenario['test_id']}")
    print(f"   Query: {scenario['user_query']}")
    print(f"   Expected: {scenario['expected_behavior']}")

print(f"\n... and {len(test_scenarios) - 5} more")

with open("output_test_scenarios.json", "w") as f:
    json.dump(test_scenarios, f, indent=2)

print("\n✓ Saved to: output_test_scenarios.json")

# ===== Final Summary =====
print("\n" + "="*80)
print("ANALYSIS COMPLETE - SUMMARY")
print("="*80)

print(f"""
✓ Analyzed {len(sheet_names)} total sheets
✓ Identified {len(sheet_names[:3])} special sheets (first 3)
✓ Identified {len(sheet_names[3:])} product sheets
✓ Created fuzzy matching solution with {len(test_queries)} test queries
✓ Generated {len(common_questions)} question categories
✓ Created {len(test_scenarios)} test scenarios

Output files created:
  1. output_sheet_names.json - Complete list of all tabs
  2. output_structure_analysis.json - Detailed structure of sheets
  3. output_fuzzy_matching_tests.json - Fuzzy matching test results
  4. output_common_questions.json - Bank of common questions
  5. output_test_scenarios.json - Comprehensive test cases

Next steps:
  1. Review the structure analysis to understand the difference between first 3 sheets and product sheets
  2. Use the fuzzy matching approach in your agent instructions
  3. Test your agent with the test scenarios
  4. Implement error handling for edge cases identified
""")

wb.close()
print("\nDone! 🎉")

